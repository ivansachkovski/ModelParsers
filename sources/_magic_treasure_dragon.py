import openpyxl
import json

def parse_excel(file_path):
    # Открытие Excel файла
    wb = openpyxl.load_workbook(file_path)

    # Выбор нужного листа
    sheet = wb['FG_MoneyBallDraw_94']

    # Определение заголовка таблицы
    expected_headers = ["Reel 1-3", "Reel 4", "Reel 5", "MPT", "Weights"]

    # Формирование результата в виде списка JSON объектов
    results = []

    # Проход по строкам и столбцам
    for row in sheet.iter_rows():
        for cell in row:
            # Поиск ячейки с заголовком таблицы
            if cell.value == expected_headers[0]:
                # Проверка, является ли следующая группа ячеек таблицей размера 5x5
                if is_valid_table(sheet, cell.row, cell.column):
                    # Извлечение данных из таблицы и формирование JSON объекта
                    json_obj = extract_table_data(sheet, cell.row, cell.column)
                    results.append(json_obj)

    # Возвращение списка JSON объектов в виде JSON массива
    return json.dumps(results, indent=4)

def is_valid_table(sheet, start_row, start_col):
    # Проверка наличия 5 строк и 5 столбцов с данными в таблице
    for i in range(start_row + 1, start_row + 6):
        for j in range(start_col, start_col + 5):
            cell = sheet.cell(row=i, column=j)
            if cell.value is None:
                return False
    return True

def extract_table_data(sheet, start_row, start_col):
    # Извлечение данных из таблицы и формирование JSON объекта
    json_obj = {'values': [], 'weights': []}
    for i in range(start_row + 1, start_row + 6):
        row_values = []
        for j in range(start_col, start_col + 3):
            row_values.append(str(sheet.cell(row=i, column=j).value))  # Получаем значение ячейки как строку
        json_obj['values'].append(','.join(row_values))
        json_obj['weights'].append(sheet.cell(row=i, column=start_col + 4).value)
    return json_obj

# Пример использования
file_path = 'file.xlsx'
result_json = parse_excel(file_path)
print(result_json)